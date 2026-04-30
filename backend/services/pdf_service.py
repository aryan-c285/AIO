"""
PDF Service — all PDF processing operations using PyMuPDF (fitz) and pikepdf.
Every function operates purely in memory via io.BytesIO.
"""

import io
from typing import List

import fitz  # PyMuPDF
import pikepdf
from docx import Document as DocxDocument
from openpyxl import Workbook


def compress_pdf(file_bytes: bytes, quality: int = 50) -> io.BytesIO:
    """
    Compress a PDF by re-saving with garbage collection and deflation.
    *quality* (1-100) controls image down-sampling — lower = smaller file.
    """
    src = fitz.open(stream=file_bytes, filetype="pdf")
    output = io.BytesIO()

    # Re-render pages at reduced resolution for images
    dpi = max(36, int(72 * quality / 100))  # scale DPI with quality

    dst = fitz.open()
    for page in src:
        # Get the page as a pixmap at reduced resolution, then re-insert
        pix = page.get_pixmap(dpi=dpi)
        img_pdf = fitz.open()
        img_page = img_pdf.new_page(width=page.rect.width, height=page.rect.height)
        img_page.insert_image(page.rect, pixmap=pix)
        dst.insert_pdf(img_pdf)
        img_pdf.close()

    dst.save(output, garbage=4, deflate=True, linear=True)
    dst.close()
    src.close()
    output.seek(0)
    return output


def merge_pdfs(files_bytes: List[bytes]) -> io.BytesIO:
    """Merge multiple PDF byte streams into a single PDF."""
    merged = fitz.open()
    for fb in files_bytes:
        doc = fitz.open(stream=fb, filetype="pdf")
        merged.insert_pdf(doc)
        doc.close()

    output = io.BytesIO()
    merged.save(output, garbage=4, deflate=True)
    merged.close()
    output.seek(0)
    return output


def split_pdf(file_bytes: bytes, page_ranges: str) -> io.BytesIO:
    """
    Split a PDF keeping only specified pages.
    *page_ranges* — comma-separated, e.g. "1-3,5,7-9".
    Returns a single PDF with the selected pages.
    """
    src = fitz.open(stream=file_bytes, filetype="pdf")
    total = src.page_count
    pages_to_keep: list[int] = []

    for part in page_ranges.split(","):
        part = part.strip()
        if "-" in part:
            start_s, end_s = part.split("-", 1)
            start = max(1, int(start_s))
            end = min(total, int(end_s))
            pages_to_keep.extend(range(start - 1, end))  # 0-indexed
        else:
            p = int(part) - 1
            if 0 <= p < total:
                pages_to_keep.append(p)

    if not pages_to_keep:
        raise ValueError("No valid pages selected")

    dst = fitz.open()
    for p in pages_to_keep:
        dst.insert_pdf(src, from_page=p, to_page=p)

    output = io.BytesIO()
    dst.save(output, garbage=4, deflate=True)
    dst.close()
    src.close()
    output.seek(0)
    return output


def pdf_to_word(file_bytes: bytes) -> io.BytesIO:
    """Extract text from each PDF page and write it into a DOCX document."""
    src = fitz.open(stream=file_bytes, filetype="pdf")
    doc = DocxDocument()

    for i, page in enumerate(src):
        if i > 0:
            doc.add_page_break()
        text = page.get_text()
        for line in text.split("\n"):
            doc.add_paragraph(line)

    src.close()
    output = io.BytesIO()
    doc.save(output)
    output.seek(0)
    return output


def pdf_to_excel(file_bytes: bytes) -> io.BytesIO:
    """
    Extract text from each PDF page and write it into an Excel workbook.
    Each page becomes a separate sheet. Lines become rows, and
    whitespace-separated tokens become columns.
    """
    src = fitz.open(stream=file_bytes, filetype="pdf")
    wb = Workbook()
    # Remove default sheet — we'll create one per page
    wb.remove(wb.active)

    for i, page in enumerate(src):
        title = f"Page {i + 1}"
        ws = wb.create_sheet(title=title)
        text = page.get_text()
        for row_idx, line in enumerate(text.split("\n"), start=1):
            tokens = line.split()
            for col_idx, token in enumerate(tokens, start=1):
                ws.cell(row=row_idx, column=col_idx, value=token)

    src.close()
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
