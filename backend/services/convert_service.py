"""
Convert Service — document format conversion using python-docx, openpyxl, and PyMuPDF.
Every function operates purely in memory via io.BytesIO.
"""

import csv
import io

import fitz  # PyMuPDF
from docx import Document as DocxDocument
from openpyxl import load_workbook


def word_to_pdf(file_bytes: bytes) -> io.BytesIO:
    """
    Convert a DOCX file to PDF.
    Extracts all text from the Word document and lays it out as PDF pages
    using PyMuPDF.  This is a text-based conversion — complex formatting,
    images, and tables may not be perfectly preserved.
    """
    doc = DocxDocument(io.BytesIO(file_bytes))

    pdf = fitz.open()
    page = pdf.new_page(width=595, height=842)  # A4 at 72 dpi
    y_cursor = 50.0
    left_margin = 50.0
    line_height = 14.0
    max_y = 792.0

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            y_cursor += line_height
            if y_cursor > max_y:
                page = pdf.new_page(width=595, height=842)
                y_cursor = 50.0
            continue

        # Wrap long lines manually (~80 chars per line at this font size)
        while text:
            chunk = text[:90]
            text = text[90:]
            if y_cursor > max_y:
                page = pdf.new_page(width=595, height=842)
                y_cursor = 50.0
            page.insert_text(
                (left_margin, y_cursor),
                chunk,
                fontsize=11,
                fontname="helv",
            )
            y_cursor += line_height

    output = io.BytesIO()
    pdf.save(output)
    pdf.close()
    output.seek(0)
    return output


def excel_to_csv(file_bytes: bytes) -> io.BytesIO:
    """
    Convert the first sheet of an XLSX workbook to a CSV file.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    output = io.BytesIO()
    text_wrapper = io.TextIOWrapper(output, encoding="utf-8", newline="")
    writer = csv.writer(text_wrapper)

    for row in ws.iter_rows(values_only=True):
        writer.writerow([cell if cell is not None else "" for cell in row])

    text_wrapper.flush()
    text_wrapper.detach()  # detach so we can seek on the underlying BytesIO
    wb.close()
    output.seek(0)
    return output
